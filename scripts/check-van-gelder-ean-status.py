#!/usr/bin/env python3
"""Van Gelder status per EAN (huidige Mima-logica).

- Bron: docs/van-gelder-mima-assortiment.csv (kolom `ean` = waarheid)
- Bestelbaar in app: actieve prijslijst (Prices API), niet ProductStatus
- ProductStatus: alleen informatief uit Articles API, opgezocht per EAN-variant
  (artikelnummer in CSV wordt alleen intern gebruikt om varianten te indexeren;
   VG heeft geen EAN-lookup endpoint)

Output: docs/van-gelder-mima-assortiment-status.xlsx
"""

from __future__ import annotations

import csv
import json
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from openpyxl import Workbook

REPO = Path(__file__).resolve().parents[1]
CSV_PATH = REPO / "docs" / "van-gelder-mima-assortiment.csv"
OUT_XLSX = REPO / "docs" / "van-gelder-mima-assortiment-status.xlsx"

IMPORT_URL = (
    "https://olcqzhxirqhkfgzgjnnw.supabase.co/functions/v1/import-van-gelder"
)
APIKEY = "sb_publishable_Xd6i1yV5VKNbYb9fz5eHyw_wRV7IYqu"


def normalize_ean(v: str) -> str:
    d = "".join(c for c in (v or "").strip() if c.isdigit())
    if len(d) == 12:
        d = "0" + d
    return d if len(d) in (13, 14) else ""


def is_orderable_product_status(product_status: str) -> bool:
    """Zelfde als dispatch-order: alleen `available` mag mee."""
    return (product_status or "").strip().lower() == "available"


def post_api(body: dict, timeout: int = 120) -> dict:
    req = urllib.request.Request(
        IMPORT_URL,
        data=json.dumps(body).encode(),
        headers={"Content-Type": "application/json", "apikey": APIKEY},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8"))


def fetch_active_price_eans() -> set[str]:
    obj = post_api({"api": "prices", "onlyActivePrices": True})
    data = obj.get("data") if isinstance(obj.get("data"), dict) else {}
    eans: set[str] = set()
    for group in data.get("Price") or []:
        if not isinstance(group, dict):
            continue
        for line in group.get("Lines") or []:
            if isinstance(line, dict) and line.get("Active") is False:
                continue
            ean = normalize_ean(str((line or {}).get("Key") or ""))
            if ean:
                eans.add(ean)
    return eans


def fetch_variants_by_artikelnummer(article_id: str) -> tuple[int | None, list[dict]]:
    """Alle EAN-varianten onder één VG-artikelnummer (niet één regel)."""
    try:
        obj = post_api({"api": "articles", "articleId": article_id}, timeout=30)
    except Exception as e:
        return None, [{"error": str(e)}]

    status = obj.get("status")
    data = obj.get("data") if isinstance(obj.get("data"), dict) else {}
    articles = data.get("Article") if isinstance(data.get("Article"), list) else []
    variants: list[dict] = []
    for a in articles:
        if not isinstance(a, dict):
            continue
        ean = normalize_ean(str(a.get("EAN") or ""))
        variants.append(
            {
                "ean": ean,
                "product_status": str(a.get("ProductStatus") or "").strip(),
                "name": str(a.get("Name") or "").strip(),
                "unit": str(a.get("UnitOfMeasurement") or "").strip(),
                "units": a.get("Units"),
                "artikelnummer": str(a.get("Id") or article_id).strip(),
            }
        )
    return int(status) if status is not None else None, variants


def build_ean_status_index(rows_in: list[dict]) -> dict[str, dict]:
    """Index ProductStatus per EAN door alle varianten per artikelnummer te laden."""
    art_ids = sorted(
        {
            (r.get("artikelnummer") or "").strip()
            for r in rows_in
            if (r.get("artikelnummer") or "").strip()
        }
    )
    print(
        f"Indexeren Articles API: {len(art_ids)} artikelnummers "
        f"(alleen intern, export is per EAN)..."
    )

    by_ean: dict[str, dict] = {}
    with ThreadPoolExecutor(max_workers=40) as pool:
        futs = {pool.submit(fetch_variants_by_artikelnummer, aid): aid for aid in art_ids}
        done = 0
        for fut in as_completed(futs):
            aid = futs[fut]
            http_status, variants = fut.result()
            done += 1
            if done % 10 == 0 or done == len(art_ids):
                print(f"  {done}/{len(art_ids)}", flush=True)
            for v in variants:
                if v.get("error"):
                    continue
                ean = v.get("ean") or ""
                if not ean:
                    continue
                # Bewaar eerste match; zelfde EAN zou zelfde status moeten hebben
                if ean not in by_ean:
                    by_ean[ean] = {
                        **v,
                        "http_status": http_status,
                        "index_artikelnummer": aid,
                    }
    return by_ean


def main() -> None:
    rows_in: list[dict] = []
    with CSV_PATH.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            rows_in.append(row)

    print("Ophalen actieve prijslijst (Prices API)...")
    active_price_eans = fetch_active_price_eans()
    print(f"  {len(active_price_eans)} EANs op prijslijst")

    by_ean = build_ean_status_index(rows_in)

    wb = Workbook()
    ws = wb.active
    ws.title = "Per_EAN"
    ws.append(
        [
            "ean",
            "naam_csv",
            "raw_ingredient_preferred",
            "product_status",
            "product_status_orderable",
            "on_active_price_list",
            "app_can_dispatch",
            "vg_variant_name",
            "vg_unit",
            "vg_units",
            "status_lookup",
            "note",
        ]
    )

    counts = {
        "rows": len(rows_in),
        "unique_ean": 0,
        "on_price_list": 0,
        "app_dispatch": 0,
        "status_found": 0,
        "status_missing": 0,
        "unavailable": 0,
    }
    seen_ean: set[str] = set()

    for row in rows_in:
        ean_raw = (row.get("ean") or "").strip()
        ean = normalize_ean(ean_raw)
        if ean and ean not in seen_ean:
            seen_ean.add(ean)
            counts["unique_ean"] += 1

        idx = by_ean.get(ean, {}) if ean else {}
        ps = idx.get("product_status") or ""
        on_list = bool(ean and ean in active_price_eans)
        app_ok = on_list and is_orderable_product_status(ps)
        if on_list:
            counts["on_price_list"] += 1
        if app_ok:
            counts["app_dispatch"] += 1

        if not ean:
            lookup = "geen_ean_in_csv"
            counts["status_missing"] += 1
        elif ean in by_ean:
            lookup = "articles_api_by_ean_variant"
            counts["status_found"] += 1
        else:
            lookup = "ean_niet_in_articles_index"
            counts["status_missing"] += 1

        if ps.lower() == "unavailable":
            counts["unavailable"] += 1

        note_parts: list[str] = []
        if ps.lower() == "inactive":
            note_parts.append("inactive → dispatch slaat regel over (alleen available mag mee)")
        if ps.lower() == "unavailable":
            note_parts.append("unavailable → dispatch slaat regel over")
        if ean and not on_list:
            note_parts.append("niet op actieve prijslijst → dispatch slaat regel over")
        if lookup == "ean_niet_in_articles_index" and on_list:
            note_parts.append(
                "wel op prijslijst; ProductStatus onbekend (geen variant in Articles-index)"
            )

        ws.append(
            [
                ean_raw or ean,
                row.get("naam", ""),
                row.get("raw_ingredient_preferred", ""),
                ps,
                "ja" if is_orderable_product_status(ps) else ("nee" if ps else ""),
                "ja" if on_list else "nee",
                "ja" if app_ok else "nee",
                idx.get("name", ""),
                idx.get("unit", ""),
                idx.get("units", ""),
                lookup,
                "; ".join(note_parts),
            ]
        )

    ws2 = wb.create_sheet("Samenvatting")
    ws2.append(["metric", "count"])
    for k, v in counts.items():
        ws2.append([k, v])
    ws2.append([])
    ws2.append(["uitleg", ""])
    ws2.append(
        [
            "product_status",
            "Informatief uit Articles API voor deze EAN-variant",
        ]
    )
    ws2.append(
        [
            "on_active_price_list / app_can_dispatch",
            "Zelfde check als live app (Prices API, onlyactiveprices=true)",
        ]
    )
    ws2.append(
        [
            "artikelnummer",
            "Niet in export; alleen gebruikt om Articles-varianten te indexeren",
        ]
    )

    # Unieke EANs
    ws3 = wb.create_sheet("Unieke_EAN")
    ws3.append(
        [
            "ean",
            "product_status",
            "on_active_price_list",
            "app_can_dispatch",
            "vg_variant_name",
            "status_lookup",
        ]
    )
    unique_rows: dict[str, dict] = {}
    for row in rows_in:
        ean = normalize_ean((row.get("ean") or "").strip())
        if ean and ean not in unique_rows:
            unique_rows[ean] = row
    for ean, row in sorted(unique_rows.items()):
        idx = by_ean.get(ean, {})
        ps = idx.get("product_status") or ""
        on_list = ean in active_price_eans
        lookup = (
            "articles_api_by_ean_variant"
            if ean in by_ean
            else "ean_niet_in_articles_index"
        )
        ws3.append(
            [
                ean,
                ps,
                "ja" if on_list else "nee",
                "ja" if on_list else "nee",
                idx.get("name", ""),
                lookup,
            ]
        )

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Klaar: {OUT_XLSX}")
    print(
        f"  regels={counts['rows']} unieke_ean={counts['unique_ean']} "
        f"status_gevonden={counts['status_found']} op_prijslijst={counts['on_price_list']} "
        f"unavailable={counts['unavailable']}"
    )


if __name__ == "__main__":
    main()
