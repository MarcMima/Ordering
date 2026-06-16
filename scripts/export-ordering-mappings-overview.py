#!/usr/bin/env python3
"""Export ordering-app mapping overview: grondstof → leverancier → VG/Bidfood codes."""

from __future__ import annotations

import json
import subprocess
from pathlib import Path

from openpyxl import Workbook

REPO = Path(__file__).resolve().parents[1]
OUT = REPO / "docs" / "ordering-app-mappings-overzicht.xlsx"

SQL = """
SELECT
  l.name AS locatie,
  r.name AS grondstof_in_app,
  s.name AS leverancier,
  si.is_preferred,
  CASE
    WHEN lower(s.name) LIKE '%van gelder%' THEN NULL
    ELSE si.supplier_article_code
  END AS artikelnummer,
  si.ean_code AS ean,
  si.supplier_article_name AS leverancier_artikelnaam,
  si.order_unit,
  si.order_unit_size,
  si.vg_is_active,
  si.vg_last_status,
  si.bf_is_active,
  sc.channel AS order_kanaal
FROM supplier_ingredients si
JOIN suppliers s ON s.id = si.supplier_id
JOIN raw_ingredients r ON r.id = si.raw_ingredient_id
JOIN locations l ON l.id = r.location_id
LEFT JOIN supplier_order_channels sc ON sc.supplier_id = s.id
WHERE EXISTS (
  SELECT 1 FROM supplier_ingredients si2
  WHERE si2.raw_ingredient_id = r.id
    AND si2.supplier_id IN (SELECT id FROM suppliers WHERE location_id = l.id)
)
ORDER BY l.name, r.name, si.is_preferred DESC, s.name;
"""

SQL_PREFERRED = """
SELECT
  l.name AS locatie,
  r.name AS grondstof_in_app,
  s.name AS leverancier,
  CASE
    WHEN lower(s.name) LIKE '%van gelder%' THEN NULL
    ELSE si.supplier_article_code
  END AS artikelnummer,
  si.ean_code AS ean,
  si.supplier_article_name AS leverancier_artikelnaam,
  si.order_unit,
  sc.channel AS order_kanaal
FROM supplier_ingredients si
JOIN suppliers s ON s.id = si.supplier_id
JOIN raw_ingredients r ON r.id = si.raw_ingredient_id
JOIN locations l ON l.id = r.location_id
LEFT JOIN supplier_order_channels sc ON sc.supplier_id = s.id
WHERE si.is_preferred = TRUE
ORDER BY l.name, s.name, r.name;
"""


def query(sql: str) -> list[dict]:
    out = subprocess.check_output(
        ["supabase", "db", "query", "--linked", "--yes", sql],
        cwd=REPO,
        text=True,
    )
    return json.loads(out[out.find("{") :])["rows"]


def main() -> None:
    all_rows = query(SQL)
    pref_rows = query(SQL_PREFERRED)

    wb = Workbook()

    ws = wb.active
    ws.title = "Preferred_bestelling"
    ws.append(
        [
            "locatie",
            "grondstof_in_app",
            "leverancier",
            "artikelnummer",
            "ean",
            "leverancier_artikelnaam",
            "order_unit",
            "order_kanaal",
        ]
    )
    for r in pref_rows:
        ws.append(
            [
                r.get("locatie"),
                r.get("grondstof_in_app"),
                r.get("leverancier"),
                r.get("artikelnummer"),
                r.get("ean"),
                r.get("leverancier_artikelnaam"),
                r.get("order_unit"),
                r.get("order_kanaal"),
            ]
        )

    ws2 = wb.create_sheet("Alle_koppelingen")
    ws2.append(
        [
            "locatie",
            "grondstof_in_app",
            "leverancier",
            "is_preferred",
            "artikelnummer",
            "ean",
            "leverancier_artikelnaam",
            "order_unit",
            "vg_is_active",
            "vg_last_status",
            "order_kanaal",
        ]
    )
    for r in all_rows:
        ws2.append(
            [
                r.get("locatie"),
                r.get("grondstof_in_app"),
                r.get("leverancier"),
                r.get("is_preferred"),
                r.get("artikelnummer"),
                r.get("ean"),
                r.get("leverancier_artikelnaam"),
                r.get("order_unit"),
                r.get("vg_is_active"),
                r.get("vg_last_status"),
                r.get("order_kanaal"),
            ]
        )

    # Duplicates: multiple preferred per raw+location
    ws3 = wb.create_sheet("Dubbel_preferred")
    ws3.append(["locatie", "grondstof_in_app", "aantal_preferred", "leveranciers"])
    from collections import defaultdict

    dup: dict[tuple[str, str], list[str]] = defaultdict(list)
    for r in pref_rows:
        key = (r.get("locatie") or "", r.get("grondstof_in_app") or "")
        dup[key].append(r.get("leverancier") or "")
    for (loc, raw), sups in sorted(dup.items()):
        if len(sups) > 1:
            ws3.append([loc, raw, len(sups), ", ".join(sups)])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"  preferred rows: {len(pref_rows)}")
    print(f"  all links: {len(all_rows)}")


if __name__ == "__main__":
    main()
