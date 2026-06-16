#!/usr/bin/env python3
"""Export Van Gelder catalog via bulk APIs (assortment + prices + article lookups).

Van Gelder design (prod):
  - Assortment  GET /api/assortment/1.0/{catalog}  → all EANs in catalog (one call)
  - Prices      GET /api/prices/1.0/all            → all price lines (one call)
  - Articles    GET /api/articles/1.0/{articleId}  → product details per VG artikelnummer

Customer MIMAMS1 has catalogs: MIMAMS1 (klantspecifiek) + HOR (hoofdassortiment ~5135 EANs).
"""

from __future__ import annotations

import json
import subprocess
import urllib.request
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook

REPO = Path(__file__).resolve().parents[1]
OUT_PATH = REPO / "docs" / "van-gelder-catalog-export.xlsx"

IMPORT_URL = (
    "https://olcqzhxirqhkfgzgjnnw.supabase.co/functions/v1/import-van-gelder"
)
APIKEY = "sb_publishable_Xd6i1yV5VKNbYb9fz5eHyw_wRV7IYqu"
ACTIVE_STATUSES = frozenset({"available", "active", "actief"})


def api_post(payload: dict) -> dict:
    req = urllib.request.Request(
        IMPORT_URL,
        data=json.dumps(payload).encode(),
        headers={"Content-Type": "application/json", "apikey": APIKEY},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=180) as resp:
        return json.loads(resp.read().decode("utf-8"))


def fetch_assortment(catalog: str) -> list[str]:
    obj = api_post({"api": "assortments", "code": catalog})
    if not obj.get("ok"):
        raise RuntimeError(f"assortment {catalog} failed: {obj}")
    products = (obj.get("data") or {}).get("Products")
    if not isinstance(products, list):
        return []
    return [str(p).strip() for p in products if p]


def fetch_prices() -> dict[str, list[dict]]:
    obj = api_post({"api": "prices", "onlyActivePrices": True})
    if not obj.get("ok"):
        raise RuntimeError(f"prices failed: {obj}")
    by_ean: dict[str, list[dict]] = defaultdict(list)
    for group in (obj.get("data") or {}).get("Price", []) or []:
        group_type = str(group.get("Type") or "")
        for line in group.get("Lines") or []:
            if not line.get("Active", True):
                continue
            key = str(line.get("Key") or "").strip()
            if not key:
                continue
            tier = (line.get("Tiers") or [{}])[0]
            by_ean[key].append(
                {
                    "price_group": group_type,
                    "customer_code": line.get("CustomerCode"),
                    "price_qty": tier.get("Quantity"),
                    "price_value": tier.get("Value"),
                    "start_date": line.get("StartDate"),
                    "end_date": line.get("EndDate"),
                }
            )
    return by_ean


def fetch_article(article_id: str) -> dict | None:
    obj = api_post({"api": "articles", "articleId": article_id})
    if obj.get("status") != 200:
        return None
    articles = (obj.get("data") or {}).get("Article")
    if not isinstance(articles, list) or not articles:
        return None
    row = articles[0]
    status = str(row.get("ProductStatus") or "").strip().lower()
    return {
        "article_id": str(row.get("Id") or article_id),
        "name": str(row.get("Name") or "").strip(),
        "ean": str(row.get("EAN") or "").strip(),
        "secondary_ean": str(row.get("SecondaryEAN") or "").strip(),
        "product_status": str(row.get("ProductStatus") or "").strip(),
        "is_active": status in ACTIVE_STATUSES,
        "unit_of_measurement": str(row.get("UnitOfMeasurement") or "").strip(),
        "units": row.get("Units"),
        "brand": str(row.get("Brand") or "").strip(),
    }


def db_article_codes() -> list[str]:
    sql = """
    SELECT DISTINCT supplier_article_code AS code
    FROM supplier_ingredients si
    JOIN suppliers s ON s.id = si.supplier_id
    WHERE lower(s.name) LIKE '%van gelder%'
      AND supplier_article_code IS NOT NULL
      AND supplier_article_code ~ '^[0-9]+$'
    ORDER BY code;
    """
    out = subprocess.check_output(
        ["supabase", "db", "query", "--linked", "--yes", sql],
        cwd=REPO,
        text=True,
    )
    obj = json.loads(out[out.find("{") :])
    return [str(r["code"]) for r in obj.get("rows", []) if r.get("code")]


def main() -> None:
    print("Fetching HOR assortment (bulk)...")
    hor_eans = fetch_assortment("HOR")
    print(f"  HOR: {len(hor_eans)} EANs")

    print("Fetching prices (bulk)...")
    prices_by_ean = fetch_prices()
    print(f"  Prices: {len(prices_by_ean)} EAN keys")

    print("Fetching article details for mapped VG artikelnummers...")
    article_by_ean: dict[str, dict] = {}
    article_by_id: dict[str, dict] = {}
    for code in db_article_codes():
        row = fetch_article(code)
        if not row:
            continue
        article_by_id[code] = row
        if row.get("ean"):
            article_by_ean[row["ean"]] = row
    print(f"  DB mappings resolved: {len(article_by_id)} artikelen")

    # Sheet 1: HOR catalog merged with prices + known article names
    wb = Workbook()
    ws = wb.active
    ws.title = "HOR_catalog"
    ws.append(
        [
            "ean",
            "in_hor_assortment",
            "has_active_price",
            "price_value",
            "price_customer_code",
            "article_id",
            "name",
            "product_status",
            "is_active",
            "unit_of_measurement",
            "units",
        ]
    )
    for ean in sorted(hor_eans):
        price_rows = prices_by_ean.get(ean, [])
        art = article_by_ean.get(ean)
        best_price = price_rows[0] if price_rows else {}
        ws.append(
            [
                ean,
                True,
                bool(price_rows),
                best_price.get("price_value"),
                best_price.get("customer_code"),
                art.get("article_id") if art else "",
                art.get("name") if art else "",
                art.get("product_status") if art else "",
                art.get("is_active") if art else "",
                art.get("unit_of_measurement") if art else "",
                art.get("units") if art else "",
            ]
        )

    # Sheet 2: all active prices
    ws2 = wb.create_sheet("Active_prices")
    ws2.append(
        [
            "ean",
            "price_group",
            "customer_code",
            "price_qty",
            "price_value",
            "start_date",
            "end_date",
        ]
    )
    for ean in sorted(prices_by_ean):
        for pr in prices_by_ean[ean]:
            ws2.append(
                [
                    ean,
                    pr.get("price_group"),
                    pr.get("customer_code"),
                    pr.get("price_qty"),
                    pr.get("price_value"),
                    pr.get("start_date"),
                    pr.get("end_date"),
                ]
            )

    # Sheet 3: article details for all DB-mapped codes (incl. inactive variants)
    ws3 = wb.create_sheet("DB_mapped_articles_live")
    ws3.append(
        [
            "article_id",
            "name",
            "ean",
            "secondary_ean",
            "product_status",
            "is_active",
            "unit_of_measurement",
            "units",
            "brand",
            "in_hor_assortment",
        ]
    )
    hor_set = set(hor_eans)
    for code in sorted(article_by_id, key=lambda c: article_by_id[c].get("name", "")):
        r = article_by_id[code]
        ws3.append(
            [
                r.get("article_id"),
                r.get("name"),
                r.get("ean"),
                r.get("secondary_ean"),
                r.get("product_status"),
                r.get("is_active"),
                r.get("unit_of_measurement"),
                r.get("units"),
                r.get("brand"),
                r.get("ean") in hor_set,
            ]
        )

    # Sheet 4: granaatappel family (quick ID range lookup)
    ws4 = wb.create_sheet("Granaatappel_varianten")
    ws4.append(
        [
            "article_id",
            "name",
            "ean",
            "product_status",
            "is_active",
            "unit_of_measurement",
            "units",
            "in_hor_assortment",
        ]
    )
    print("Scanning granaatappel artikelnummer-range 166100-166250...")
    for aid in range(166100, 166251):
        row = fetch_article(str(aid))
        if not row:
            continue
        if "granaat" not in row.get("name", "").lower():
            continue
        ws4.append(
            [
                row.get("article_id"),
                row.get("name"),
                row.get("ean"),
                row.get("product_status"),
                row.get("is_active"),
                row.get("unit_of_measurement"),
                row.get("units"),
                row.get("ean") in hor_set,
            ]
        )

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
