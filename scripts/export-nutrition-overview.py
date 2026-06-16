#!/usr/bin/env python3
"""Export voedingswaarden naar Excel: gerechten, grondstoffen, halffabrikaten."""

from __future__ import annotations

import json
import subprocess
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[1]
OUT = REPO / "docs" / "mima-voedingswaarden-overzicht.xlsx"

SQL_GERECHTEN = """
SELECT
  mi.name AS gerecht,
  mi.category AS categorie,
  mi.subcategory AS subcategorie,
  mi.active AS actief,
  minu.kcal AS decl_kcal,
  minu.protein_g AS decl_eiwit_g,
  minu.carbs_g AS decl_koolhydraten_g,
  minu.sugar_g AS decl_suikers_g,
  minu.fat_g AS decl_vet_g,
  minu.sat_fat_g AS decl_verzadigd_vet_g,
  minu.fiber_g AS decl_vezel_g,
  minu.salt_g AS decl_zout_g,
  minu.source AS decl_bron,
  cmn.kcal AS berek_kcal,
  cmn.protein_g AS berek_eiwit_g,
  cmn.carbs_g AS berek_koolhydraten_g,
  cmn.sugar_g AS berek_suikers_g,
  cmn.fat_g AS berek_vet_g,
  cmn.sat_fat_g AS berek_verzadigd_vet_g,
  cmn.fiber_g AS berek_vezel_g,
  cmn.salt_g AS berek_zout_g,
  cmn.has_lab_inputs AS berek_heeft_lab_input,
  cmn.missing_inputs AS berek_mist_input,
  CASE
    WHEN minu.kcal IS NULL OR cmn.kcal IS NULL OR minu.kcal = 0 THEN NULL
    ELSE ROUND(((cmn.kcal - minu.kcal) / minu.kcal) * 100.0, 1)
  END AS kcal_verschil_pct
FROM menu_items mi
LEFT JOIN menu_item_nutrition minu ON minu.menu_item_id = mi.id
LEFT JOIN computed_menu_item_nutrition cmn ON cmn.menu_item_id = mi.id
ORDER BY mi.category, mi.display_order, mi.name;
"""

SQL_GRONDSTOFFEN = """
SELECT DISTINCT ON (lower(btrim(r.name)))
  r.name AS grondstof,
  r.unit AS eenheid,
  inv.kcal_per_100g AS kcal_per_100g,
  inv.protein_g AS eiwit_g,
  inv.carbs_g AS koolhydraten_g,
  inv.sugar_g AS suikers_g,
  inv.fat_g AS vet_g,
  inv.sat_fat_g AS verzadigd_vet_g,
  inv.fiber_g AS vezel_g,
  inv.salt_g AS zout_g,
  inv.source AS bron,
  inv.source_type AS bron_type,
  inv.measured_at AS gemeten_op,
  inv.verified_by AS geverifieerd_door
FROM raw_ingredients r
LEFT JOIN ingredient_nutritional_values inv ON inv.raw_ingredient_id = r.id
ORDER BY lower(btrim(r.name)), inv.kcal_per_100g DESC NULLS LAST;
"""

SQL_HALFFABRIKATEN = """
SELECT
  pi.name AS halffabrikaat,
  pi.unit AS telt_eenheid,
  pi.content_amount AS inhoud_hoeveelheid,
  cpn.kcal_per_100g,
  cpn.protein_per_100g AS eiwit_per_100g,
  cpn.carbs_per_100g AS koolhydraten_per_100g,
  cpn.sugar_per_100g AS suikers_per_100g,
  cpn.fat_per_100g AS vet_per_100g,
  cpn.sat_fat_per_100g AS verzadigd_vet_per_100g,
  cpn.fiber_per_100g AS vezel_per_100g,
  cpn.salt_per_100g AS zout_per_100g,
  ppn.kcal AS kcal_per_eenheid,
  ppn.protein_g AS eiwit_per_eenheid,
  ppn.carbs_g AS koolhydraten_per_eenheid,
  ppn.sugar_g AS suikers_per_eenheid,
  ppn.fat_g AS vet_per_eenheid,
  ppn.sat_fat_g AS verzadigd_vet_per_eenheid,
  ppn.fiber_g AS vezel_per_eenheid,
  ppn.salt_g AS zout_per_eenheid,
  coalesce(ppn.source, 'berekend uit recept') AS bron,
  ppn.source_type AS bron_type,
  cpn.has_lab_inputs AS heeft_lab_input,
  cpn.missing_nutrition_inputs AS mist_recept_input
FROM prep_items pi
LEFT JOIN computed_prep_item_nutrition cpn ON cpn.prep_item_id = pi.id
LEFT JOIN prep_item_nutritional_values ppn ON ppn.prep_item_id = pi.id
ORDER BY pi.name;
"""


def query(sql: str) -> list[dict]:
    out = subprocess.check_output(
        ["supabase", "db", "query", "--linked", "--yes", sql],
        cwd=REPO,
        text=True,
    )
    return json.loads(out[out.find("{") :])["rows"]


def autosize(ws) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 48)


def write_sheet(ws, headers: list[str], rows: list[dict], keys: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(k) for k in keys])
    autosize(ws)


def main() -> None:
    gerechten = query(SQL_GERECHTEN)
    grondstoffen = query(SQL_GRONDSTOFFEN)
    halffabrikaten = query(SQL_HALFFABRIKATEN)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Gerechten"
    write_sheet(
        ws1,
        [
            "Gerecht",
            "Categorie",
            "Subcategorie",
            "Actief",
            "Decl. kcal (portie)",
            "Decl. eiwit (g)",
            "Decl. koolhydraten (g)",
            "Decl. suikers (g)",
            "Decl. vet (g)",
            "Decl. verzadigd vet (g)",
            "Decl. vezel (g)",
            "Decl. zout (g)",
            "Decl. bron",
            "Berek. kcal (portie)",
            "Berek. eiwit (g)",
            "Berek. koolhydraten (g)",
            "Berek. suikers (g)",
            "Berek. vet (g)",
            "Berek. verzadigd vet (g)",
            "Berek. vezel (g)",
            "Berek. zout (g)",
            "Berek. lab-input",
            "Berek. mist input",
            "Kcal verschil %",
        ],
        gerechten,
        [
            "gerecht",
            "categorie",
            "subcategorie",
            "actief",
            "decl_kcal",
            "decl_eiwit_g",
            "decl_koolhydraten_g",
            "decl_suikers_g",
            "decl_vet_g",
            "decl_verzadigd_vet_g",
            "decl_vezel_g",
            "decl_zout_g",
            "decl_bron",
            "berek_kcal",
            "berek_eiwit_g",
            "berek_koolhydraten_g",
            "berek_suikers_g",
            "berek_vet_g",
            "berek_verzadigd_vet_g",
            "berek_vezel_g",
            "berek_zout_g",
            "berek_heeft_lab_input",
            "berek_mist_input",
            "kcal_verschil_pct",
        ],
    )

    ws2 = wb.create_sheet("Grondstoffen")
    write_sheet(
        ws2,
        [
            "Grondstof",
            "Eenheid",
            "kcal per 100g/ml",
            "Eiwit (g)",
            "Koolhydraten (g)",
            "Suikers (g)",
            "Vet (g)",
            "Verzadigd vet (g)",
            "Vezel (g)",
            "Zout (g)",
            "Bron",
            "Bron type",
            "Gemeten op",
            "Geverifieerd door",
        ],
        grondstoffen,
        [
            "grondstof",
            "eenheid",
            "kcal_per_100g",
            "eiwit_g",
            "koolhydraten_g",
            "suikers_g",
            "vet_g",
            "verzadigd_vet_g",
            "vezel_g",
            "zout_g",
            "bron",
            "bron_type",
            "gemeten_op",
            "geverifieerd_door",
        ],
    )

    ws3 = wb.create_sheet("Halffabrikaten")
    write_sheet(
        ws3,
        [
            "Halffabrikaat",
            "Telt-eenheid",
            "Inhoud hoeveelheid",
            "kcal per 100g",
            "Eiwit per 100g (g)",
            "Koolhydraten per 100g (g)",
            "Suikers per 100g (g)",
            "Vet per 100g (g)",
            "Verzadigd vet per 100g (g)",
            "Vezel per 100g (g)",
            "Zout per 100g (g)",
            "kcal per eenheid",
            "Eiwit per eenheid (g)",
            "Koolhydraten per eenheid (g)",
            "Suikers per eenheid (g)",
            "Vet per eenheid (g)",
            "Verzadigd vet per eenheid (g)",
            "Vezel per eenheid (g)",
            "Zout per eenheid (g)",
            "Bron",
            "Bron type",
            "Lab-input",
            "Mist recept-input",
        ],
        halffabrikaten,
        [
            "halffabrikaat",
            "telt_eenheid",
            "inhoud_hoeveelheid",
            "kcal_per_100g",
            "eiwit_per_100g",
            "koolhydraten_per_100g",
            "suikers_per_100g",
            "vet_per_100g",
            "verzadigd_vet_per_100g",
            "vezel_per_100g",
            "zout_per_100g",
            "kcal_per_eenheid",
            "eiwit_per_eenheid",
            "koolhydraten_per_eenheid",
            "suikers_per_eenheid",
            "vet_per_eenheid",
            "verzadigd_vet_per_eenheid",
            "vezel_per_eenheid",
            "zout_per_eenheid",
            "bron",
            "bron_type",
            "heeft_lab_input",
            "mist_recept_input",
        ],
    )

    ws4 = wb.create_sheet("Info")
    ws4.append(["MIMA voedingswaarden export"])
    ws4.append(["Gegenereerd op", date.today().isoformat()])
    ws4.append([])
    ws4.append(["Tab Gerechten", "Officiële waarden (menu_item_nutrition) + berekende som uit componenten per portie"])
    ws4.append(["Tab Grondstoffen", "Per 100 g/ml; één rij per unieke grondstofnaam"])
    ws4.append(["Tab Halffabrikaten", "Per 100 g (berekend/lab) + per telt-eenheid waar beschikbaar"])
    autosize(ws4)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"  gerechten: {len(gerechten)}")
    print(f"  grondstoffen: {len(grondstoffen)}")
    print(f"  halffabrikaten: {len(halffabrikaten)}")


if __name__ == "__main__":
    main()
